import asyncio
from playwright.async_api import async_playwright
from agent_visualizer.visualizer import AgentVisualizer


async def run_agent():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()

        # Step 1: Open Google
        await page.goto("https://www.google.com")

        await page.fill("textarea[name='q']", "Playwright Python tutorial")
        await page.keyboard.press("Enter")

        await page.wait_for_timeout(2000)

        await page.click("h3")

        await page.wait_for_timeout(3000)
        await browser.close()


if __name__ == "__main__":
    with AgentVisualizer() as visualizer:
        print("Live View URL:", visualizer.get_live_view_url())
        asyncio.run(run_agent())


#Excel
"""
Excel router – project-based generation.

GET /api/v1/excel/generate/{project_id}
    1. Queries processeddata WHERE project_id = :id AND resolution_status = 'completed'
    2. Merges all output JSON blobs into one ExcelSchema
    3. Streams back a .xlsx file

GET /api/v1/excel/sample/xlsx   – download a pre-filled sample .xlsx
GET /api/v1/excel/sample/csv    – download a sample template .csv
"""

import logging
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
import io

from src.application.services.excel_builder_service import (
    ExcelBuilderService,
    ExcelSchema,
    QAColumn,
    QARecord,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/excel", tags=["excel"])

# ── Service singleton ─────────────────────────────────────────────────────────

_service_instance = ExcelBuilderService()

def get_excel_service() -> ExcelBuilderService:
    return _service_instance

# ── DB dependency ─────────────────────────────────────────────────────────────

def get_db():
    from src.infrastructure.database import SessionLocal
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ── Response helpers ──────────────────────────────────────────────────────────

def _excel_response(excel_bytes: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_response(csv_bytes: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Schema builder ────────────────────────────────────────────────────────────

def _frame_schema(rows: list) -> ExcelSchema:
    """
    Build an ExcelSchema from the `output` column of each DB row.

    Expected output JSON shape per row:
    {
        "qa_columns": [                       ← taken from first row only
            {
                "section": "...",
                "attribute_code": "A1",
                "attribute_name": "...",
                "sub_attribute": "..."
            }
        ],
        "records": [                          ← flat dict, one per row
            { "A1": "Y", "A2": "N/A", ... }
        ],
        "metadata": { ... }                   ← optional, taken from first row
    }
    """
    def _output(row) -> dict:
        raw = row.output if hasattr(row, "output") else row.get("output", {})
        return raw if isinstance(raw, dict) else {}

    first_output = _output(rows[0])

    # qa_columns come from the first row only
    raw_qa_cols = first_output.get("qa_columns", [])
    if not raw_qa_cols:
        raise ValueError("No qa_columns found in output JSON.")
    qa_columns = [QAColumn(**col) for col in raw_qa_cols]

    # One QARecord per DB row – the record is the first entry in "records"
    records: list[QARecord] = []
    for row in rows:
        output     = _output(row)
        row_records = output.get("records", [])
        qa_results  = row_records[0] if row_records else {}
        records.append(QARecord(qa_results=qa_results))

    metadata = first_output.get("metadata", {})

    return ExcelSchema(metadata=metadata, qa_columns=qa_columns, records=records)


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get(
    "/generate/{project_id}",
    summary="Generate QA Review Excel for a project",
    status_code=status.HTTP_200_OK,
)
def generate_excel_from_db(
    project_id: str,
    db=Depends(get_db),
    service: ExcelBuilderService = Depends(get_excel_service),
):
    logger.info("Excel generation requested | project_id=%s", project_id)

    # 1. Pull completed rows
    try:
        from src.infrastructure.models import ProcessedData

        rows = (
            db.query(ProcessedData)
            .filter(
                ProcessedData.project_id == project_id,
                ProcessedData.resolution_status == "completed",
            )
            .order_by(ProcessedData.sample_number)
            .all()
        )
    except Exception as exc:
        logger.exception("DB query failed | project_id=%s", project_id)
        raise HTTPException(status_code=500, detail=f"DB query failed: {exc}")

    if not rows:
        raise HTTPException(
            status_code=404,
            detail=f"No completed records found for project_id={project_id!r}.",
        )

    logger.info("Rows fetched | project_id=%s, count=%d", project_id, len(rows))

    # 2. Build schema
    try:
        schema = _frame_schema(rows)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    # 3. Generate Excel
    try:
        excel_bytes = service.build_excel(schema)
    except Exception as exc:
        logger.exception("Excel build failed | project_id=%s", project_id)
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {exc}")

    production_month = schema.metadata.get("production_month") or project_id
    filename = f"QA_Review_{production_month}.xlsx"
    logger.info("Excel ready | file=%s, size=%d bytes", filename, len(excel_bytes))
    return _excel_response(excel_bytes, filename)


@router.get(
    "/sample/xlsx",
    summary="Download a sample QA Review Excel file",
    status_code=status.HTTP_200_OK,
)
def download_sample_excel(service: ExcelBuilderService = Depends(get_excel_service)):
    excel_bytes = service.build_sample_excel()
    return _excel_response(excel_bytes, "QA_Review_Sample.xlsx")


@router.get(
    "/sample/csv",
    summary="Download a sample QA results CSV template",
    status_code=status.HTTP_200_OK,
)
def download_sample_csv(service: ExcelBuilderService = Depends(get_excel_service)):
    """
    Returns a CSV with:
      Row 1 (header): attribute codes   A1, A2, A3, B1, ...
      Row 2 (example): sample values    Y,  N/A, N/A, Y, ...
    """
    csv_bytes = service.build_sample_csv()
    return _csv_response(csv_bytes, "QA_Review_Template.csv")


# ── Health ────────────────────────────────────────────────────────────────────

@router.get("/health", include_in_schema=False, status_code=200)
def health():
    return {"service": "excel", "status": "ok",
            "timestamp": datetime.now(timezone.utc).isoformat()}
#excel builder serviuce 
"""
Class-based stateless service for QA Review Excel generation.
Summary sheet is only written when metadata is non-empty.
Columns are driven entirely by qa_columns from the schema.
"""

import csv
import io
from itertools import groupby
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Palette ───────────────────────────────────────────────────────────────────
DARK_HEADER  = "1F3864"
MID_HEADER   = "2F5496"
LIGHT_HEADER = "D6E4F7"
YELLOW_INPUT = "FFFF00"
GREY_LEGEND  = "D9D9D9"
WHITE        = "FFFFFF"
ALT_ROW      = "EBF3FB"


# ── Dataclass models ──────────────────────────────────────────────────────────
@dataclass
class QAColumn:
    section: str
    attribute_code: str
    attribute_name: str
    sub_attribute: str


@dataclass
class QARecord:
    qa_results: dict = field(default_factory=dict)


@dataclass
class ExcelSchema:
    metadata: dict[str, Any]
    qa_columns: list[QAColumn]
    records: list[QARecord]

    @classmethod
    def from_dict(cls, data: dict) -> "ExcelSchema":
        return cls(
            metadata=data.get("metadata", {}),
            qa_columns=[QAColumn(**c) for c in data.get("qa_columns", [])],
            records=[
                QARecord(qa_results=r)
                for r in data.get("records", [])
            ],
        )


# ── Service ───────────────────────────────────────────────────────────────────
class ExcelBuilderService:

    def build_excel(self, schema: ExcelSchema) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        if schema.metadata:
            self._build_summary_sheet(wb, schema.metadata)

        self._build_status_sheet(wb, schema)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_sample_excel(self) -> bytes:
        schema = ExcelSchema.from_dict(self._sample_schema())
        return self.build_excel(schema)

    def build_sample_csv(self) -> bytes:
        """
        Returns a CSV whose header row contains each attribute_code and
        whose single data row shows example Y / N / N/A values.

        Format:
            A1,A2,A3,...
            Y,N/A,N/A,...
        """
        sample = self._sample_schema()
        qa_cols = sample["qa_columns"]
        records = sample["records"]

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([c["attribute_code"] for c in qa_cols])
        for rec in records:
            writer.writerow([rec.get(c["attribute_code"], "") for c in qa_cols])

        return buf.getvalue().encode("utf-8")

    def get_sample_schema(self) -> dict:
        return self._sample_schema()

    # ── Sheet: Summary ────────────────────────────────────────────────────────

    def _build_summary_sheet(self, wb: Workbook, meta: dict):
        ws = wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 60

        meta_rows = [
            ("QA Review Program",      meta.get("qa_review_program",      "")),
            ("Review being Performed",  meta.get("review_being_performed", "")),
            ("Month Testing Performed", meta.get("month_testing_performed","")),
            ("Consultant",              meta.get("consultant",             "")),
            ("Production Month",        meta.get("production_month",       "")),
            ("Population Size",         meta.get("population_size",        "")),
            ("Sample Size",             meta.get("sample_size",            "")),
            ("Sample Size Methodology", meta.get("sample_size_methodology","")),
        ]

        filled_rows = [(label, value) for label, value in meta_rows if value != ""]
        if not filled_rows:
            return

        for i, (label, value) in enumerate(filled_rows, start=1):
            lbl = ws.cell(row=i, column=2, value=label)
            val = ws.cell(row=i, column=3, value=value)
            lbl.font      = self._font(bold=True, size=10)
            lbl.fill      = self._fill(GREY_LEGEND)
            lbl.alignment = self._align(h="right")
            lbl.border    = self._border()
            val.font      = self._font(size=10)
            val.fill      = self._fill(GREY_LEGEND)
            val.alignment = self._align(h="left")
            val.border    = self._border()
            ws.row_dimensions[i].height = 28 if label == "Sample Size Methodology" else 18

        legend_start = len(filled_rows) + 3
        legend_items = [
            ("Testing Legend", "",     True),
            ("Y =", "Yes – Attribute Satisfied; No exception noted.", False),
            ("N =", "No – Attribute NOT Satisfied; Exception noted.", False),
            ("NA",  "Test procedure not applicable to sample.",       False),
        ]
        for j, (code, desc, is_header) in enumerate(legend_items, start=legend_start):
            c = ws.cell(row=j, column=2, value=code)
            d = ws.cell(row=j, column=3, value=desc)
            for cell in (c, d):
                cell.fill   = self._fill(GREY_LEGEND)
                cell.border = self._border()
                cell.font   = self._font(bold=is_header, size=10)
            if is_header:
                ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=3)
                c.alignment = self._align(h="center")
            else:
                c.alignment = self._align(h="center")
                d.alignment = self._align(h="left")
            ws.row_dimensions[j].height = 18

    # ── Sheet: Event Closed Status ────────────────────────────────────────────

    def _build_status_sheet(self, wb: Workbook, schema: ExcelSchema):
        ws = wb.create_sheet("Event Closed Status")
        ws.freeze_panes = "A5"
        ws.sheet_view.showGridLines = False

        qa_cols = schema.qa_columns
        n_cols  = len(qa_cols)

        for i in range(n_cols):
            ws.column_dimensions[get_column_letter(i + 1)].width = 16

        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 70
        ws.row_dimensions[3].height = 90
        ws.row_dimensions[4].height = 14

        self._write_section_row(ws, qa_cols)
        self._write_attribute_row(ws, qa_cols)
        self._write_subattribute_row(ws, qa_cols)

        # Row 4: thin filter/spacer row
        for col in range(1, n_cols + 1):
            c = ws.cell(row=4, column=col, value="")
            c.fill   = self._fill(LIGHT_HEADER)
            c.border = self._border()

        for row_idx, rec in enumerate(schema.records, start=5):
            self._write_data_row(ws, rec, qa_cols, row_idx)

        ws.auto_filter.ref = f"A4:{get_column_letter(n_cols)}4"

    def _write_section_row(self, ws, qa_cols: list[QAColumn]):
        col_cursor = 1
        for section_name, group in self._group_sections(qa_cols):
            span = len(group)
            self._merge_cell(
                ws, 1, col_cursor, 1, col_cursor + span - 1,
                section_name, DARK_HEADER, "FFFFFF", bold=True,
            )
            col_cursor += span

    def _write_attribute_row(self, ws, qa_cols: list[QAColumn]):
        for i, col in enumerate(qa_cols, start=1):
            label = f"{col.attribute_code}: {col.attribute_name}"
            self._header_cell(ws, 2, i, label, MID_HEADER, size=8)

    def _write_subattribute_row(self, ws, qa_cols: list[QAColumn]):
        for i, col in enumerate(qa_cols, start=1):
            c = ws.cell(row=3, column=i, value=col.sub_attribute)
            c.font      = self._font(size=8)
            c.fill      = self._fill(YELLOW_INPUT)
            c.alignment = self._align()
            c.border    = self._border()

    def _write_data_row(self, ws, rec: QARecord, qa_cols: list[QAColumn], row_idx: int):
        row_fill = ALT_ROW if row_idx % 2 == 0 else WHITE
        for i, col_def in enumerate(qa_cols, start=1):
            val = rec.qa_results.get(col_def.attribute_code, "")
            c = ws.cell(row=row_idx, column=i, value=val)
            c.font      = self._font(size=9)
            c.fill      = self._fill(YELLOW_INPUT if val else row_fill)
            c.alignment = self._align()
            c.border    = self._border()

    # ── Style helpers ─────────────────────────────────────────────────────────

    @staticmethod
    def _font(bold=False, color="000000", size=9, name="Arial") -> Font:
        return Font(bold=bold, color=color, size=size, name=name)

    @staticmethod
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    @staticmethod
    def _align(h="center", v="center", wrap=True) -> Alignment:
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    @staticmethod
    def _border() -> Border:
        t = Side(border_style="thin", color="000000")
        return Border(left=t, right=t, top=t, bottom=t)

    def _header_cell(self, ws, row, col, value, bg,
                     font_color="FFFFFF", bold=True, size=9):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = self._font(bold=bold, color=font_color, size=size)
        c.fill      = self._fill(bg)
        c.alignment = self._align()
        c.border    = self._border()
        return c

    def _merge_cell(self, ws, r1, c1, r2, c2, value,
                    bg, font_color="000000", bold=False, size=10):
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1,
                           end_row=r2,   end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        cell.font      = self._font(bold=bold, color=font_color, size=size)
        cell.fill      = self._fill(bg)
        cell.alignment = self._align()
        cell.border    = self._border()
        return cell

    @staticmethod
    def _group_sections(qa_cols: list[QAColumn]):
        result = []
        for key, group in groupby(qa_cols, key=lambda x: x.section):
            result.append((key, list(group)))
        return result

    @staticmethod
    def _sample_schema() -> dict:
        return {
            "metadata": {
                "qa_review_program":      "Event Driven Reviews (EDR)",
                "review_being_performed": "Event Closed",
                "month_testing_performed":"January",
                "consultant":             "Firas Faraj & Keishia Lindsey",
                "production_month":       "December",
                "population_size":        1670,
                "sample_size":            145,
                "sample_size_methodology": (
                    "Utilizes the Wells Fargo Independent Risk Management (IRM) "
                    "Sample Size Calculator to determine sample sizes when conducting reviews."
                ),
            },
            "qa_columns": [
                {
                    "section": "Event Creation",
                    "attribute_code": "A1",
                    "attribute_name": "Validate Event Description is complete",
                    "sub_attribute": "A1a – Validate the event description is complete (for purposes of review)",
                },
                {
                    "section": "Event Creation",
                    "attribute_code": "A2",
                    "attribute_name": "Validate Recommendation Rationale aligns with Recommended Action",
                    "sub_attribute": "A2a – If EDR Event type is not Priority FC Business Referral, validate the recommendation rationale aligns with risk action recommendation type",
                },
                {
                    "section": "Event Creation",
                    "attribute_code": "A3",
                    "attribute_name": "Validate Recommendation Rationale aligns with Event Description",
                    "sub_attribute": "A3a – If EDR event type is not Priority FC Business Referral, validate the recommendation rationale aligns with the event description",
                },
                {
                    "section": "Level 1 Review (Priority FC Business)",
                    "attribute_code": "B1",
                    "attribute_name": "Validate Recommendation Rationale aligns with Recommended Action",
                    "sub_attribute": "B1a – Validate the recommendation rationale aligns with risk action recommendation type",
                },
                {
                    "section": "Level 1 Review (Priority FC Business)",
                    "attribute_code": "B2",
                    "attribute_name": "Validate Recommendation Rationale aligns with Event Description",
                    "sub_attribute": "B2a – Validate the recommendation rationale is reasonable based on the details of the event description",
                },
                {
                    "section": "Level 2 Review",
                    "attribute_code": "C1",
                    "attribute_name": "Validate the decision rationale aligns with decision type",
                    "sub_attribute": "C1a – Validate the decision rationale aligns with the risk decision type",
                },
                {
                    "section": "Level 2 Review",
                    "attribute_code": "C2",
                    "attribute_name": "Validate the decision rationale aligns with Event Description and Recommended Action Rationale",
                    "sub_attribute": "C2a – Validate the decision rationale aligns with Event Description and Recommended Action Rationale",
                },
                {
                    "section": "Level 2 Review",
                    "attribute_code": "C3",
                    "attribute_name": "Validate Level 2 Rationale does not list incorrect group that made recommendation",
                    "sub_attribute": "C3a – Based on the above grid, either ensure no reference is made to recommending group or the correct group is referenced in rationale, based on Event Type and Brokerage field",
                },
                {
                    "section": "Level 2 Review",
                    "attribute_code": "C4",
                    "attribute_name": "Validate if all the Elevated Approvals 1 obtained per requirements",
                    "sub_attribute": "C4a – Confirm if approver decision and rationale were obtained for all Elevated Approvals 1",
                },
                {
                    "section": "Initiation of Decision Execution",
                    "attribute_code": "D1",
                    "attribute_name": "Validate if the profile refresh request was referred to the appropriate process",
                    "sub_attribute": "D1a – Is the profile refresh referred to the appropriate group and evidence provided to QA",
                },
                {
                    "section": "Initiation of Decision Execution",
                    "attribute_code": "D2",
                    "attribute_name": "Validate if the Risk Rating Override request was referred to the appropriate process",
                    "sub_attribute": "D2a – Is a risk rating override request present in the risk rating override appropriate group and Actimize workflow for this event (Actimize only)",
                },
                {
                    "section": "Initiation of Decision Execution",
                    "attribute_code": "D3",
                    "attribute_name": "Validate if customer exit request was referred to the appropriate process",
                    "sub_attribute": "D3a – Is the Exit referred to the appropriate group and evidence provided to QA",
                },
            ],
            "records": [
                {
                    "A1": "Y",   "A2": "N/A", "A3": "N/A",
                    "B1": "Y",   "B2": "Y",
                    "C1": "N/A", "C2": "N/A", "C3": "N/A", "C4": "N/A",
                    "D1": "N/A", "D2": "N/A", "D3": "N/A",
                },
            ],
        }